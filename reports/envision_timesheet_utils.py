"""
Envision GEO — Timesheet Excel Report Generator
Produces an .xlsx file grouped by user -> week (Sun-Sat) with one row per
calendar day (all of that day's time entries accumulated together) showing
RT/OT/Meals, plus a weekly total and a grand total per user.

RT/OT policy (Envision):
  - Alberta statutory holiday: the full day worked is Overtime, and none of
    it counts toward the week's 44h Regular allowance.
  - Otherwise: first 8h of each day = Regular candidate; anything beyond
    8h/day = OT. The week's Regular candidates are capped at 44h (Sun-Sat);
    once the running weekly total hits 44h, further regular-candidate hours
    also become OT. Hours already OT from the daily rule are never
    re-tested against the weekly cap (no double counting).
duration field is stored in minutes.
"""

import io
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.utils.envision_time import alberta_stat_holidays, utc_to_envision_local

# ── Style constants ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
TOTAL_FILL  = PatternFill("solid", fgColor="DCE6F1")
GRAND_FILL  = PatternFill("solid", fgColor="1F3864")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
GRAND_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

THIN        = Side(style="thin",   color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS    = ["Employee Name", "Week", "Date & Day", "Reg Hrs", "OT Hours", "Total Hours", "Meals"]
COL_WIDTHS = [22, 8, 28, 10, 10, 12, 8]

DAILY_RT_HOURS  = 8
WEEKLY_RT_HOURS = 44


# ── Helpers ────────────────────────────────────────────────────────────────────

def _clean_hrs(v):
    """Return int for whole numbers, rounded float otherwise. None if zero."""
    if not v:
        return None
    r = round(v, 2)
    return int(r) if r == int(r) else r


def _week_sunday(d):
    """Return the Sunday that starts the week containing d (week = Sun-Sat)."""
    return d - timedelta(days=(d.weekday() + 1) % 7)


def _split_day_hours(day_total_hours, cumulative_reg_before_today, is_stat_holiday=False):
    """
    Envision RT/OT for one calendar day, aware of the running weekly total:
      0. Alberta stat holiday: the full day is OT, and it doesn't touch the
         week's Regular allowance at all.
      1. Otherwise, first 8h of the day = Regular candidate; beyond 8h/day
         = OT.
      2. The Regular-candidate pool is capped at 44h for the week — once the
         running weekly total hits 44h, further regular-candidate hours also
         become OT (daily-OT hours are never re-tested against the cap).
    Returns (day_reg, day_ot, new_cumulative_reg).
    """
    if is_stat_holiday:
        return 0.0, day_total_hours, cumulative_reg_before_today

    daily_reg_candidate = min(day_total_hours, DAILY_RT_HOURS)
    daily_ot = max(0.0, day_total_hours - DAILY_RT_HOURS)

    remaining_weekly_allowance = max(0.0, WEEKLY_RT_HOURS - cumulative_reg_before_today)
    day_reg = min(daily_reg_candidate, remaining_weekly_allowance)
    day_ot = daily_ot + (daily_reg_candidate - day_reg)

    return day_reg, day_ot, cumulative_reg_before_today + day_reg


def _cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ── Main builder ───────────────────────────────────────────────────────────────

def generate_envision_timesheet_xlsx(
    entries_by_user,
    project_name="",
    task_name="",
    date_from=None,
    date_to=None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Row 1 — title
    ws.merge_cells("A1:G1")
    ws["A1"] = "Envision GEO - Timesheet Report"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    # Row 2 — project
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Project:  {project_name or '-'}"
    ws["A2"].font      = Font(name="Calibri", bold=True, size=11)
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    # Row 3 — task
    ws.merge_cells("A3:G3")
    ws["A3"] = f"Task:     {task_name or '-'}"
    ws["A3"].font      = Font(name="Calibri", bold=True, size=11)
    ws["A3"].alignment = LEFT
    ws.row_dimensions[3].height = 18

    # Row 4 — date range
    ws.merge_cells("A4:G4")
    if date_from and date_to:
        ws["A4"] = "Period:   {} - {}".format(
            date_from.strftime("%b %d, %Y"),
            date_to.strftime("%b %d, %Y"),
        )
    ws["A4"].font      = Font(name="Calibri", italic=True, size=10, color="595959")
    ws["A4"].alignment = LEFT
    ws.row_dimensions[4].height = 16

    # Row 5 — spacer
    ws.row_dimensions[5].height = 6

    # Row 6 — column headers
    for col_idx, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        _cell(ws, 6, col_idx, header,
              font=HEADER_FONT, fill=HEADER_FILL,
              align=CENTER, border=THIN_BORDER)
        ws.column_dimensions[ws.cell(6, col_idx).column_letter].width = width
    ws.row_dimensions[6].height = 18

    current_row = 7

    # ── Alberta stat holidays for the years this report spans ─────────────────
    if date_from and date_to:
        holiday_years = range(date_from.year, date_to.year + 1)
    else:
        all_dates = [
            utc_to_envision_local(entry.start_time).date()
            for entries in entries_by_user.values()
            for entry in entries
        ]
        holiday_years = range(min(d.year for d in all_dates), max(d.year for d in all_dates) + 1) if all_dates else []
    stat_holidays = alberta_stat_holidays(holiday_years)

    # ── Per-user data ──────────────────────────────────────────────────────────
    for user_name, entries in entries_by_user.items():

        # Accumulate every entry into its Mountain-Time calendar day first —
        # multiple entries on the same day become ONE row, not repeated rows.
        days = defaultdict(lambda: {"minutes": 0, "meals": False})
        for entry in entries:
            day = utc_to_envision_local(entry.start_time).date()
            days[day]["minutes"] += entry.duration or 0
            if entry.meals:
                days[day]["meals"] = True

        # Group those days into weeks (Sun-Sat)
        weeks = defaultdict(list)
        for day in days:
            weeks[_week_sunday(day)].append(day)

        week_number = 1
        user_rt     = 0.0
        user_ot     = 0.0
        user_total  = 0.0
        user_meals  = 0

        for week_sun in sorted(weeks.keys()):
            week_days  = sorted(weeks[week_sun])
            week_label = "Week{}".format(week_number)

            week_rt    = 0.0
            week_ot    = 0.0
            week_total = 0.0
            week_meals = 0
            cumulative_reg = 0.0  # the 44h regular cap resets every week

            for day in week_days:
                day_info = days[day]
                day_total_hours = day_info["minutes"] / 60.0
                rt, ot, cumulative_reg = _split_day_hours(
                    day_total_hours, cumulative_reg, is_stat_holiday=day in stat_holidays
                )
                total     = rt + ot
                meals_val = 1 if day_info["meals"] else None
                day_str   = day.strftime("%A, %d-%m-%Y")

                week_rt    += rt
                week_ot    += ot
                week_total += total
                if day_info["meals"]:
                    week_meals += 1

                user_rt    += rt
                user_ot    += ot
                user_total += total
                if day_info["meals"]:
                    user_meals += 1

                row_data = [
                    user_name, week_label, day_str,
                    _clean_hrs(rt), _clean_hrs(ot), _clean_hrs(total), meals_val,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    _cell(ws, current_row, col_idx, value,
                          font=NORMAL_FONT,
                          align=LEFT if col_idx == 1 else CENTER,
                          border=THIN_BORDER)
                ws.row_dimensions[current_row].height = 16
                current_row += 1

            # Weekly total row
            week_total_data = [
                "Total Hours", None, None,
                _clean_hrs(week_rt),
                _clean_hrs(week_ot),
                _clean_hrs(week_total),
                week_meals if week_meals else None,
            ]
            for col_idx, value in enumerate(week_total_data, start=1):
                _cell(ws, current_row, col_idx, value,
                      font=BOLD_FONT, fill=TOTAL_FILL,
                      align=LEFT if col_idx == 1 else CENTER,
                      border=THIN_BORDER)
            ws.row_dimensions[current_row].height = 16
            current_row += 1
            week_number += 1

        # Employee grand total row (all weeks combined)
        grand_data = [
            "{} - Total".format(user_name), None, None,
            _clean_hrs(user_rt),
            _clean_hrs(user_ot),
            _clean_hrs(user_total),
            user_meals if user_meals else None,
        ]
        for col_idx, value in enumerate(grand_data, start=1):
            _cell(ws, current_row, col_idx, value,
                  font=GRAND_FONT, fill=GRAND_FILL,
                  align=LEFT if col_idx == 1 else CENTER,
                  border=THIN_BORDER)
        ws.row_dimensions[current_row].height = 17
        current_row += 1

        # Blank separator between users
        current_row += 1

    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

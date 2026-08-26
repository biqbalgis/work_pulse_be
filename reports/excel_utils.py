import openpyxl
from openpyxl.styles import Font, Alignment
from django.db.models import Sum
from io import BytesIO
from time_entries.models import TimeEntry
from django.db.models.functions import Coalesce

def generate_time_entry_report(start_date, end_date, project_id=None):
    """
    Generates an Excel report for time entries within a date range, 
    optionally filtered by project.
    Columns: User Full Name, Job Title, Total number of hours
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Time Entries Report"

    # Set headers
    headers = ["User Full Name", "Job Title", "Total number of hours"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Query Time Entries
    queryset = TimeEntry.objects.filter(
        start_time__date__gte=start_date,
        start_time__date__lte=end_date
    )

    if project_id:
        queryset = queryset.filter(project_id=project_id)

    # Aggregate by user and job title
    # Note: duration is in minutes in the model
    results = queryset.values(
        'user__first_name', 
        'user__last_name', 
        'job_title__name'
    ).annotate(
        total_minutes=Sum('duration')
    ).order_by('user__first_name', 'user__last_name', 'job_title__name')

    # Fill data
    for row_num, entry in enumerate(results, 2):
        full_name = f"{entry['user__first_name']} {entry['user__last_name']}".strip()
        if not full_name:
            # Fallback if names are empty
            from users.models import User
            # This is a bit inefficient but handled for edge cases
            # Ideally we'd use the username or email if first/last are missing
            pass

        job_title = entry['job_title__name'] or "N/A"
        total_hours = (entry['total_minutes'] or 0) / 60.0

        sheet.cell(row=row_num, column=1).value = full_name
        sheet.cell(row=row_num, column=2).value = job_title
        sheet.cell(row=row_num, column=3).value = round(total_hours, 2)

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

"""
Envision GEO — Costing LEM Excel Generator
Same data contract and layout as envision_costing_pdf_utils.generate_envision_costing_pdf —
see that module's docstring for the full `data` dict shape.

Table columns (matches the PDF exactly):
  Name/Asset | Date | Work Type | Task | Description | Hrs/Units | Rate | Total

Layout:
  - Header block: Project / Task / Job Number / Client on the left,
    PM / Contact / Phone on the right, LEM No. / Date on the far right.
  - Labour rows grouped by the linked asset's name ONLY (job titles play no
    part in this report) → one subtotal row per asset.
  - Grand Total row. There is no separate Asset section — every asset usage
    is folded into the labour rows above.
  - Footer: Client Rep + signature line.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Style constants (colour-matched to the PDF) ───────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="7F7F7F")   # HEADER_GRAY in the PDF
SUBTOTAL_FILL = PatternFill("solid", fgColor="DCE6F1")   # SUBTOTAL_BG in the PDF
GRAND_FILL    = PatternFill("solid", fgColor="1F3864")   # GRAND_BG in the PDF

TITLE_FONT   = Font(name="Calibri", bold=True, size=13, color="1F3864")
LABEL_FONT   = Font(name="Calibri", bold=True, size=10)
VALUE_FONT   = Font(name="Calibri", size=10)
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
GRAND_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

THIN        = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS    = ["Name", "Date", "Work Type", "Task", "Description", "Hrs / Units", "Rate", "Total"]
COL_WIDTHS = [22, 13, 14, 14, 32, 12, 12, 14]

MONEY_COLS = {7, 8}   # 1-indexed: Rate, Total


def _fmt(v):
    """Return a clean number string — int if whole, otherwise 2dp."""
    try:
        f = float(str(v).replace(",", "").replace("$", ""))
        return int(f) if f == int(f) else round(f, 2)
    except (ValueError, TypeError):
        return v or ""


def _cell(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


def _header_field(ws, row, label_col, label, value_col, value):
    _cell(ws, row, label_col, label, font=LABEL_FONT, align=LEFT)
    _cell(ws, row, value_col, value, font=VALUE_FONT, align=LEFT)


def generate_envision_costing_xlsx(data: dict) -> io.BytesIO:
    """Build the Envision Costing LEM Excel workbook from `data` and return a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Costing LEM"

    for idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "Envision GEO - Costing LEM"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    # ── Header info block (Project/Task/Job Number/Client | PM/Contact/Phone | LEM/Date) ──
    _header_field(ws, 2, 1, "Project:", 2, data.get("project_name", ""))
    _header_field(ws, 2, 5, "PM:", 6, data.get("pm_name", ""))
    _header_field(ws, 2, 7, "LEM No.:", 8, data.get("lem_number", ""))

    _header_field(ws, 3, 1, "Task:", 2, data.get("task_name", ""))
    _header_field(ws, 3, 5, "Contact:", 6, data.get("pm_contact", ""))
    _header_field(ws, 3, 7, "Date:", 8, data.get("lem_date", ""))

    _header_field(ws, 4, 1, "Job Number:", 2, data.get("job_number", ""))
    _header_field(ws, 4, 5, "Phone No.:", 6, data.get("pm_phone", ""))

    _header_field(ws, 5, 1, "Client:", 2, data.get("client", ""))

    for r in range(2, 6):
        ws.row_dimensions[r].height = 16

    # ── Spacer ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 8

    # ── Column headers ─────────────────────────────────────────────────────────
    header_row = 7
    for col_idx, header in enumerate(COLUMNS, start=1):
        _cell(ws, header_row, col_idx, header,
              font=HEADER_FONT, fill=HEADER_FILL, align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[header_row].height = 18

    current_row = header_row + 1

    def _write_row(values, font=NORMAL_FONT, fill=None):
        nonlocal current_row
        for col_idx, value in enumerate(values, start=1):
            align = LEFT if col_idx in (1, 5) else (RIGHT if col_idx in MONEY_COLS else CENTER)
            number_format = "$#,##0.00" if col_idx in MONEY_COLS and value not in (None, "") else None
            _cell(
                ws, current_row, col_idx, value,
                font=font, fill=fill, align=align, border=THIN_BORDER,
                number_format=number_format,
            )
        current_row += 1

    # ── Labour groups (one per Work Type, all employees combined) ────────────────
    for group in data.get("labour_groups", []):
        job_title = group.get("job_title", "")

        for entry in group.get("entries", []):
            _write_row([
                entry.get("employee", ""),
                entry.get("date", ""),
                job_title,
                entry.get("task", ""),
                entry.get("description", ""),
                _fmt(entry.get("hours", "")),
                _fmt(entry.get("rate", "")),
                _fmt(entry.get("total", "")),
            ])

        _write_row(
            ["", "", "", "", "{} Total".format(job_title),
             _fmt(group.get("hours_total", "0")), "", _fmt(group.get("subtotal", "0"))],
            font=BOLD_FONT, fill=SUBTOTAL_FILL,
        )

    # ── Asset rows ─────────────────────────────────────────────────────────────
    for asset in data.get("asset_rows", []):
        _write_row([
            asset.get("name", ""),
            asset.get("date", ""),
            "",
            "",
            "",
            _fmt(asset.get("hours_units", "")),
            _fmt(asset.get("rate", "")),
            _fmt(asset.get("total", "")),
        ])

    if data.get("asset_rows"):
        _write_row(
            ["", "", "", "", "", "", "Asset Total", _fmt(data.get("asset_total", "0"))],
            font=BOLD_FONT, fill=SUBTOTAL_FILL,
        )

    # ── Grand total ────────────────────────────────────────────────────────────
    _write_row(
        ["", "", "", "", "", "", "Total", _fmt(data.get("grand_total", "0"))],
        font=GRAND_FONT, fill=GRAND_FILL,
    )

    # ── Footer: client rep + signature ────────────────────────────────────────
    current_row += 1
    _cell(ws, current_row, 1, "Client Rep: {}".format(data.get("client_rep", "")), font=NORMAL_FONT, align=LEFT)
    current_row += 1
    if data.get("sign"):
        _cell(
            ws, current_row, 1,
            "Signed: {}    Date: {}".format(data.get("sign_name", ""), data.get("sign_date", "")),
            font=NORMAL_FONT, align=LEFT,
        )
    else:
        _cell(ws, current_row, 1, "Signature:", font=NORMAL_FONT, align=LEFT)

    ws.freeze_panes = "A{}".format(header_row + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
